import os, sys, math
from datetime import date, datetime
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import pandas as pd, yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

SECTORS = {
    "Tech & Innovation": [("XLK", 0.5), ("SOXX", 0.5)],
    "Energy & Materials": [("XLE", 0.5), ("XME", 0.5)],
    "Precious Metals": [("GLD", 0.6), ("GDX", 0.4)],
    "Crypto & Digital Assets": [("BTC-USD", 0.6), ("ETH-USD", 0.4)],
    "Fixed Income": [("AGG", 0.7), ("IEF", 0.3)],
    "Consumer": [("XLY", 0.6), ("XLP", 0.4)],
    "Health & Biotech": [("XLV", 0.6), ("XBI", 0.4)],
    "Real Assets": [("VNQ", 0.6), ("XLU", 0.4)],
    "Agriculture": [("DBA", 1.0)],
    "Cash / Liquidity": [("BIL", 1.0)],
}
HEADER = ["Year-Month", "Notes"] + list(SECTORS.keys()) + ["Top #1", "Top #2", "Average Return %"]
EXCEL_FILE = "Field_Elevate_Sector_Monthly_Returns.xlsx"

def prev_month(today): 
    first = date(today.year, today.month, 1)
    prev = first - relativedelta(days=1)
    return prev.year, prev.month

def month_bounds(y, m):
    end = monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, end)

def last_close(ticker, y, m):
    start, end = month_bounds(y, m)
    end_plus = end + relativedelta(days=1)
    try:
        df = yf.download(ticker, start=start, end=end_plus, progress=False)
        if df.empty: return None
        col = "Adj Close" if "Adj Close" in df.columns else "Close"
        return float(df[col].dropna().iloc[-1].item())
    except: return None

def monthly_return(ticker, y, m):
    lc = last_close(ticker, y, m)
    py, pm = (y, m-1) if m > 1 else (y-1, 12)
    prev = last_close(ticker, py, pm)
    return None if not lc or not prev else (lc/prev - 1)

def weighted_return(y, m, basket):
    total, wsum = 0, 0
    for t, w in basket:
        r = monthly_return(t, y, m)
        if r is not None: total += w*r; wsum += w
    return None if wsum==0 else total/wsum

def ensure_excel(path):
    if os.path.exists(path): return
    wb = Workbook(); ws = wb.active; ws.title = "Monthly"
    ws.append(HEADER)
    for c in range(1, len(HEADER)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0056A3")
    wb.save(path)

def write_data(wb, ym, data):
    ws = wb.active
    rows = {str(ws.cell(r,1).value): r for r in range(2, ws.max_row+1)}
    row = rows.get(ym) or ws.max_row+1
    if row == ws.max_row+1: ws.append([ym, ""])
    vals = []
    for i, s in enumerate(SECTORS):
        v = data.get(s); ws.cell(row, 3+i, value=None if v is None else round(v*100,2))
        vals.append(v if v is not None else -1e9)
    try:
        top1 = max(range(len(vals)), key=lambda i: vals[i])
        tmp = vals[top1]; vals[top1] = -1e10
        top2 = max(range(len(vals)), key=lambda i: vals[i]); vals[top1] = tmp
        ws.cell(row, 3+len(SECTORS), value=list(SECTORS)[top1])
        ws.cell(row, 4+len(SECTORS), value=list(SECTORS)[top2])
    except: pass
    clean = [v for v in data.values() if v is not None]
    ws.cell(row, 5+len(SECTORS), value=round(sum(clean)/len(clean)*100,2) if clean else None)

def main():
    y,m = prev_month(datetime.utcnow().date())
    ym = f"{y}-{m:02d}"
    sector_ret = {s: weighted_return(y,m,b) for s,b in SECTORS.items()}
    ensure_excel(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    write_data(wb, ym, sector_ret)
    wb.save(EXCEL_FILE)
    print(f"[OK] Updated {EXCEL_FILE} for {ym}")
    for s,v in sector_ret.items(): print(f"{s:22s}: {'N/A' if v is None else f'{v*100:.2f}%'}")

if __name__ == "__main__":
    try: main()
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr); sys.exit(1)
