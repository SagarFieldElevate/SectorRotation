#!/usr/bin/env python3
"""
Backfill historical sector returns into Excel file.
Usage:
    python backfill_historical_data.py [start_year] [start_month]

Examples:
    python backfill_historical_data.py 2017 1    # Start from Jan 2017
    python backfill_historical_data.py 2020 1    # Start from Jan 2020
    python backfill_historical_data.py           # Default: Jan 2017
"""

import os
import sys
import math
from datetime import date, datetime
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# Import from the main script
SECTORS = {
    "Tech & Innovation": [("XLK", 0.5), ("SOXX", 0.5)],
    "Energy & Materials": [("XLE", 0.5), ("XME", 0.5)],
    "Precious Metals": [("GLD", 0.6), ("GDX", 0.4)],
    "Crypto & Digital Assets": [("BTC-USD", 0.6), ("ETH-USD", 0.4)],
    "Fixed Income": [("AGG", 0.7), ("IEF", 0.3)],
    "Consumer": [("XLY", 0.6), ("XLP", 0.4)],
    "Health & Biotech": [("XLV", 0.6), ("XBI", 0.4)],
    "Real Assets": [("VNQ", 0.6), ("XLU", 0.4)],
    "Agriculture": [("DBA", 1.0)],
    "Cash / Liquidity": [("BIL", 1.0)],
}
HEADER = ["Year-Month", "Notes"] + list(SECTORS.keys()) + ["Top #1", "Top #2", "Average Return %"]
EXCEL_FILE = "Field_Elevate_Sector_Monthly_Returns.xlsx"

def month_bounds(y, m):
    end = monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, end)

def last_close(ticker, y, m):
    """Get last close price for a ticker in a given month."""
    start, end = month_bounds(y, m)
    end_plus = end + relativedelta(days=1)
    try:
        df = yf.download(ticker, start=start, end=end_plus, progress=False)
        if df.empty:
            return None
        col = "Adj Close" if "Adj Close" in df.columns else "Close"
        return float(df[col].dropna().iloc[-1].item())
    except Exception as e:
        print(f"    Warning: Could not fetch {ticker} for {y}-{m:02d}: {e}")
        return None

def monthly_return(ticker, y, m):
    """Calculate monthly return for a ticker."""
    lc = last_close(ticker, y, m)
    py, pm = (y, m-1) if m > 1 else (y-1, 12)
    prev = last_close(ticker, py, pm)
    return None if not lc or not prev else (lc/prev - 1)

def weighted_return(y, m, basket):
    """Calculate weighted return for a basket of tickers."""
    total, wsum = 0, 0
    for t, w in basket:
        r = monthly_return(t, y, m)
        if r is not None:
            total += w*r
            wsum += w
    return None if wsum==0 else total/wsum

def ensure_excel(path):
    """Ensure Excel file exists with proper headers."""
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws.append(HEADER)
    for c in range(1, len(HEADER)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0056A3")
    wb.save(path)

def write_data(wb, ym, data):
    """Write monthly data to Excel workbook."""
    ws = wb.active
    rows = {str(ws.cell(r,1).value): r for r in range(2, ws.max_row+1)}
    row = rows.get(ym) or ws.max_row+1

    # Create new row if needed
    if row == ws.max_row+1:
        ws.append([ym, ""])

    # Write sector returns
    vals = []
    for i, s in enumerate(SECTORS):
        v = data.get(s)
        ws.cell(row, 3+i, value=None if v is None else round(v*100, 2))
        vals.append(v if v is not None else -1e9)

    # Calculate top performers
    try:
        top1 = max(range(len(vals)), key=lambda i: vals[i])
        tmp = vals[top1]
        vals[top1] = -1e10
        top2 = max(range(len(vals)), key=lambda i: vals[i])
        vals[top1] = tmp
        ws.cell(row, 3+len(SECTORS), value=list(SECTORS)[top1])
        ws.cell(row, 4+len(SECTORS), value=list(SECTORS)[top2])
    except:
        pass

    # Calculate average return
    clean = [v for v in data.values() if v is not None]
    ws.cell(row, 5+len(SECTORS), value=round(sum(clean)/len(clean)*100, 2) if clean else None)

def backfill(start_year, start_month, end_year=None, end_month=None):
    """
    Backfill historical data from start_year/start_month to end_year/end_month.
    If end is not specified, goes up to last complete month.
    """
    ensure_excel(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)

    # Determine end date (default to previous month)
    if end_year is None or end_month is None:
        today = datetime.utcnow().date()
        first = date(today.year, today.month, 1)
        prev = first - relativedelta(days=1)
        end_year, end_month = prev.year, prev.month

    print(f"Backfilling data from {start_year}-{start_month:02d} to {end_year}-{end_month:02d}")
    print("=" * 80)

    current_year, current_month = start_year, start_month
    months_processed = 0

    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
        ym = f"{current_year}-{current_month:02d}"
        print(f"\nProcessing {ym}...")

        # Calculate sector returns
        sector_ret = {}
        for s, b in SECTORS.items():
            ret = weighted_return(current_year, current_month, b)
            sector_ret[s] = ret
            status = "N/A" if ret is None else f"{ret*100:6.2f}%"
            print(f"  {s:25s}: {status}")

        # Write to Excel
        write_data(wb, ym, sector_ret)
        months_processed += 1

        # Move to next month
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    # Save workbook
    wb.save(EXCEL_FILE)
    print("\n" + "=" * 80)
    print(f"✓ Successfully backfilled {months_processed} months of data")
    print(f"✓ Saved to {EXCEL_FILE}")

def main():
    # Parse command line arguments
    if len(sys.argv) >= 3:
        start_year = int(sys.argv[1])
        start_month = int(sys.argv[2])
    else:
        # Default to January 2017 (when ETH-USD data becomes available)
        start_year = 2017
        start_month = 1
        print(f"No start date specified, using default: {start_year}-{start_month:02d}")
        print(f"To specify a different start: python {sys.argv[0]} YEAR MONTH")
        print(f"Example: python {sys.argv[0]} 2020 1")
        print()

    # Validate inputs
    if not (1 <= start_month <= 12):
        print(f"Error: Invalid month {start_month}. Must be 1-12.")
        sys.exit(1)

    if start_year < 1990 or start_year > datetime.utcnow().year:
        print(f"Error: Invalid year {start_year}.")
        sys.exit(1)

    # Run backfill
    try:
        backfill(start_year, start_month)
    except Exception as e:
        print(f"\n[ERROR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
