#!/usr/bin/env python3
"""
Backfill historical sector returns using Twelve Data API.

Setup:
    1. Install: pip install twelvedata
    2. Set API key: export TWELVEDATA_API_KEY="your_api_key"
    3. Or pass as argument: python backfill_twelvedata.py 2017 1 --api-key YOUR_KEY

Usage:
    python backfill_twelvedata.py [start_year] [start_month] [--api-key KEY]

Examples:
    python backfill_twelvedata.py 2017 1
    python backfill_twelvedata.py 2020 1 --api-key abc123
"""

import os
import sys
import math
from datetime import date, datetime
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

try:
    from twelvedata import TDClient
except ImportError:
    print("Error: twelvedata package not installed.")
    print("Install with: pip install twelvedata")
    sys.exit(1)

# Import from the main script
SECTORS = {
    "Tech & Innovation": [("XLK", 0.5), ("SOXX", 0.5)],
    "Energy & Materials": [("XLE", 0.5), ("XME", 0.5)],
    "Precious Metals": [("GLD", 0.6), ("GDX", 0.4)],
    "Crypto & Digital Assets": [("BTC/USD", 0.6), ("ETH/USD", 0.4)],  # Note: format change for Twelve Data
    "Fixed Income": [("AGG", 0.7), ("IEF", 0.3)],
    "Consumer": [("XLY", 0.6), ("XLP", 0.4)],
    "Health & Biotech": [("XLV", 0.6), ("XBI", 0.4)],
    "Real Assets": [("VNQ", 0.6), ("XLU", 0.4)],
    "Agriculture": [("DBA", 1.0)],
    "Cash / Liquidity": [("BIL", 1.0)],
}
HEADER = ["Year-Month", "Notes"] + list(SECTORS.keys()) + ["Top #1", "Top #2", "Average Return %"]
EXCEL_FILE = "Field_Elevate_Sector_Monthly_Returns.xlsx"

# Global client
td_client = None

def init_client(api_key=None):
    """Initialize Twelve Data client."""
    global td_client
    if api_key is None:
        api_key = os.environ.get("TWELVEDATA_API_KEY")
    if not api_key:
        print("Error: Twelve Data API key not found.")
        print("Set environment variable: export TWELVEDATA_API_KEY='your_key'")
        print("Or pass as argument: --api-key YOUR_KEY")
        sys.exit(1)
    td_client = TDClient(apikey=api_key)

def month_bounds(y, m):
    end = monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, end)

def last_close_td(ticker, y, m):
    """Get last close price for a ticker in a given month using Twelve Data."""
    start, end = month_bounds(y, m)
    try:
        # Determine exchange based on ticker type
        exchange = None
        interval = "1day"

        # Crypto needs different handling
        if "/" in ticker:  # Crypto pair
            symbol = ticker
        else:
            symbol = ticker

        # Fetch time series data
        ts = td_client.time_series(
            symbol=symbol,
            interval=interval,
            start_date=start.strftime("%Y-%m-%d"),
            end_date=end.strftime("%Y-%m-%d"),
            outputsize=31  # Max days in a month
        )

        df = ts.as_pandas()
        if df.empty:
            return None

        # Get last close price
        return float(df['close'].iloc[-1])
    except Exception as e:
        print(f"    Warning: Could not fetch {ticker} for {y}-{m:02d}: {e}")
        return None

def monthly_return_td(ticker, y, m):
    """Calculate monthly return for a ticker using Twelve Data."""
    lc = last_close_td(ticker, y, m)
    py, pm = (y, m-1) if m > 1 else (y-1, 12)
    prev = last_close_td(ticker, py, pm)
    return None if not lc or not prev else (lc/prev - 1)

def weighted_return(y, m, basket):
    """Calculate weighted return for a basket of tickers."""
    total, wsum = 0, 0
    for t, w in basket:
        r = monthly_return_td(t, y, m)
        if r is not None:
            total += w*r
            wsum += w
    return None if wsum==0 else total/wsum

def ensure_excel(path):
    """Ensure Excel file exists with proper headers."""
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws.append(HEADER)
    for c in range(1, len(HEADER)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0056A3")
    wb.save(path)

def write_data(wb, ym, data):
    """Write monthly data to Excel workbook."""
    ws = wb.active
    rows = {str(ws.cell(r,1).value): r for r in range(2, ws.max_row+1)}
    row = rows.get(ym) or ws.max_row+1

    if row == ws.max_row+1:
        ws.append([ym, ""])

    vals = []
    for i, s in enumerate(SECTORS):
        v = data.get(s)
        ws.cell(row, 3+i, value=None if v is None else round(v*100, 2))
        vals.append(v if v is not None else -1e9)

    try:
        top1 = max(range(len(vals)), key=lambda i: vals[i])
        tmp = vals[top1]
        vals[top1] = -1e10
        top2 = max(range(len(vals)), key=lambda i: vals[i])
        vals[top1] = tmp
        ws.cell(row, 3+len(SECTORS), value=list(SECTORS)[top1])
        ws.cell(row, 4+len(SECTORS), value=list(SECTORS)[top2])
    except:
        pass

    clean = [v for v in data.values() if v is not None]
    ws.cell(row, 5+len(SECTORS), value=round(sum(clean)/len(clean)*100, 2) if clean else None)

def backfill(start_year, start_month, end_year=None, end_month=None):
    """Backfill historical data using Twelve Data API."""
    ensure_excel(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)

    if end_year is None or end_month is None:
        today = datetime.utcnow().date()
        first = date(today.year, today.month, 1)
        prev = first - relativedelta(days=1)
        end_year, end_month = prev.year, prev.month

    print(f"Backfilling data from {start_year}-{start_month:02d} to {end_year}-{end_month:02d}")
    print("Using Twelve Data API")
    print("=" * 80)

    current_year, current_month = start_year, start_month
    months_processed = 0

    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
        ym = f"{current_year}-{current_month:02d}"
        print(f"\nProcessing {ym}...")

        sector_ret = {}
        for s, b in SECTORS.items():
            ret = weighted_return(current_year, current_month, b)
            sector_ret[s] = ret
            status = "N/A" if ret is None else f"{ret*100:6.2f}%"
            print(f"  {s:25s}: {status}")

        write_data(wb, ym, sector_ret)
        months_processed += 1

        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    wb.save(EXCEL_FILE)
    print("\n" + "=" * 80)
    print(f"✓ Successfully backfilled {months_processed} months of data")
    print(f"✓ Saved to {EXCEL_FILE}")

def main():
    # Parse command line arguments
    api_key = None
    args = sys.argv[1:]

    # Extract API key if provided
    if "--api-key" in args:
        idx = args.index("--api-key")
        api_key = args[idx + 1]
        args = args[:idx] + args[idx+2:]

    if len(args) >= 2:
        start_year = int(args[0])
        start_month = int(args[1])
    else:
        start_year = 2017
        start_month = 1
        print(f"No start date specified, using default: {start_year}-{start_month:02d}")
        print(f"To specify: python {sys.argv[0]} YEAR MONTH [--api-key KEY]")
        print()

    # Validate inputs
    if not (1 <= start_month <= 12):
        print(f"Error: Invalid month {start_month}. Must be 1-12.")
        sys.exit(1)

    if start_year < 1990 or start_year > datetime.utcnow().year:
        print(f"Error: Invalid year {start_year}.")
        sys.exit(1)

    # Initialize Twelve Data client
    init_client(api_key)

    # Run backfill
    try:
        backfill(start_year, start_month)
    except Exception as e:
        print(f"\n[ERROR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
